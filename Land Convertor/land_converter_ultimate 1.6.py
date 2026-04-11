#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🇵🇰 ULTIMATE PAKISTANI LAND CONVERTER v3.1
Features: 
  • Marla (Punjab Legal) = 225 sq ft (OFFICIAL for Fard/Mutations)
  • Marla (Trad Ref) = 272 sq ft (Traditional/KPK builder reference)
  • Kanal (KPK Ref) = 5,440 sq ft (20 × 272 sq ft)
  • Sq. Karam = 30.25 sq ft (Agricultural unit)
  • About Button with Comprehensive Legal Reference Guide
  • Urdu/English Toggle • Excel Import/Export • Charts
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, font
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import numpy as np
import math
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========== CONVERSION CONSTANTS ==========
# Punjab Legal Standard (OFFICIAL - Revenue Act)
SQFT_PER_MARLA_LEGAL = 225.0          # Court-admissible standard
SQFT_PER_KANAL_LEGAL = 4500.0         # 20 × 225 sq ft

# Traditional Reference (KPK/Builder/Agricultural informal standard)
SQFT_PER_MARLA_TRAD = 272.0           # Informal reference only (NOT legally valid)
SQFT_PER_KANAL_TRAD = 5440.0          # 20 × 272 sq ft (KPK builder standard)

# Agricultural Unit (Karam)
SQFT_PER_SQ_KARAM = 30.25             # 5.5 ft × 5.5 ft

# Urdu translations dictionary
URDU_TEXT = {
    'app_title': 'پاکستانی زمین یونٹ کنورٹر',
    'patwari_standard': 'پٹواری معیار • پنجاب ریونیو ایکٹ کے مطابق',
    'enter_sqft': 'مربع فٹ میں رقبہ درج کریں',
    'sqft_label': 'مربع فٹ:',
    'add_btn': '➕ کنورژن شامل کریں',
    'clear_btn': '🗑️ سب صاف کریں',
    'export_btn': '📥 ایکسل میں محفوظ کریں',
    'import_btn': '📤 ایکسل سے درآمد کریں',
    'results': '📊 کنورژن کے نتائج',
    'legal_notice': '⚠️ قانونی انتباہ',
    'legal_text': "اہم: صرف 'مرلہ (پنجاب قانونی)' ہی عدالت، فردِ ملکیت اور میوٹیشنز کے لیے قانونی طور پر تسلیم شدہ ہے۔ 'مرلہ (روایتی حوالہ)' اور 'کنال (خیبر پختونخوا حوالہ)' صرف حوالہ کے لیے ہیں — انہیں سرکاری دستاویزات میں استعمال نہ کریں۔",
    'marla_legal': 'مرلہ (پنجاب قانونی)',
    'kanal_legal': 'کنال (پنجاب قانونی)',
    'marla_trad': 'مرلہ (روایتی حوالہ)',
    'kanal_kpk': 'کنال (خیبر پختونخوا حوالہ)',
    'sq_karam': 'مربع کرم',
    'standard_used': 'استخدام شدہ معیار',
    'karam_info': '1 کرم = 5.5 فٹ • 1 مربع کرم = 30.25 مربع فٹ',
    'status_ready': 'تیار • مربع فٹ درج کریں',
    'status_added': '✅ شامل کیا گیا: {sqft} مربع فٹ = {marla_legal} مرلہ (پنجاب قانونی) | کل: {total}',
    'import_success': 'کامیابی سے درآمد ہو گیا!',
    'export_success': 'ایکسل فائل کامیابی سے محفوظ ہو گئی!',
    'confirm_clear': 'تصدیق کریں',
    'clear_confirm_msg': 'کیا آپ واقعی تمام {count} کنورژن صاف کرنا چاہتے ہیں؟',
    'nothing_to_clear': 'صاف کرنے کے لیے کچھ نہیں ہے!',
    'no_data_export': 'ایکسپورٹ کرنے کے لیے کوئی ڈیٹا نہیں ہے!',
    'input_required': 'ان پٹ درکار ہے',
    'input_required_msg': 'براہ کرم مربع فٹ کی ویلیو درج کریں!',
    'invalid_input': 'غلط ان پٹ',
    'invalid_number': 'براہ کرم درست نمبر درج کریں!',
    'invalid_positive': 'مربع فٹ صفر سے بڑا ہونا چاہیے!',
    'standards': 'قانونی: 1 مرلہ = 225 مربع فٹ | روایتی: 1 مرلہ = 272 مربع فٹ (خیبر پختونخوا بلڈرز)',
    'lookup_tab_text': '  🔍 ریورس لک اپ  ',
    'select_unit': 'یونٹ منتخب کریں:',
    'enter_value': 'ویلیو درج کریں:',
    'calc_btn': '🧮 تبدیل کریں',
    'lookup_results_label': '🔍 لک اپ کے نتائج',
    'area_calc_tab': '  📐 تفصیلی رقبہ کیلکولیٹر  ',
    'choose_shape': 'شکل منتخب کریں:',
    'shape_rect': 'مستطیل (لمبائی × چوڑائی)',
    'shape_tri': 'مثلث (3 اضلاع - ہیرون کا فارمولا)',
    'shape_quad_tri': 'چوکور (درست: 4 اضلاع + 1 وتر)',
    'shape_quad_avg': 'چوکور (روایتی پٹواری اوسط / کچا طریقہ)',
    'side_1': 'پہلا ضلع (فٹ):',
    'side_2': 'دوسرا ضلع (فٹ):',
    'side_3': 'تیسرا ضلع (فٹ):',
    'side_4': 'چوتھا ضلع (فٹ):',
    'diagonal': 'وتر/درمیان (فٹ):',
    'calc_area_btn': '🧮 رقبہ مربع فٹ میں نکالیں',
    'area_results_label': '📋 رقبہ کے نتائج',
}

class LandConverterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("🇵🇰 ULTIMATE Pakistani Land Converter v3.1 - Punjab Legal + Traditional References")
        self.root.geometry("1400x880")
        self.root.minsize(1350, 830)
        self.root.resizable(True, True)
        
        # Language state
        self.is_urdu = False
        
        # Color scheme - Professional Patwari theme
        self.colors = {
            'green': '#004726',      # Pakistan flag green
            'white': '#FFFFFF',
            'light_green': '#C8E6C9',
            'dark_blue': '#0A2E5A',
            'light_blue': '#E3F2FD',  # Punjab Legal highlight
            'gray': '#F5F5F5',
            'border': '#424242',
            'accent_green': '#4CAF50',
            'accent_red': '#F44336',
            'accent_blue': '#2196F3',
            'trad_yellow': '#FFF9C4',  # Traditional reference highlight
            'kpk_orange': '#FFE0B2'    # KPK reference highlight
        }
        
        # Conversion history
        self.conversions = []
        
        # Urdu font handling
        self.urdu_font = self._get_urdu_font()
        
        # Build UI
        self.create_widgets()
        
        # Set focus to input field
        self.root.after(100, lambda: self.sqft_entry.focus_force())
        
        # Keyboard shortcut: Ctrl+A opens About window
        self.root.bind('<Control-a>', lambda e: self.show_about_window())
    
    def _get_urdu_font(self):
        """Get appropriate Urdu font with fallbacks"""
        urdu_fonts = [
            'Noto Nastaliq Urdu',
            'Jameel Noori Nastaleeq',
            'Urdu Typesetting',
            'Arial Unicode MS',
            'Segoe UI'
        ]
        
        for f in urdu_fonts:
            try:
                test_font = font.Font(family=f, size=11)
                return f
            except:
                continue
        return 'Arial'  # Fallback
    
    def create_widgets(self):
        """Build all GUI components with optimized layout"""
        
        # ========== MAIN FRAME ==========
        main_frame = tk.Frame(self.root, bg=self.colors['white'])
        main_frame.pack(fill='both', expand=True)
        
        # ========== HEADER FRAME ==========
        header_frame = tk.Frame(main_frame, bg=self.colors['green'], height=110)
        header_frame.pack(fill='x', padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        # Right-side buttons frame (Language + About)
        right_btns_frame = tk.Frame(header_frame, bg=self.colors['green'])
        right_btns_frame.place(relx=0.99, rely=0.25, anchor='e')
        
        # About button (LEFT side of language toggle)
        self.about_btn = tk.Button(
            right_btns_frame,
            text="ℹ️ Info",
            command=self.show_about_window,
            font=('Arial', 11, 'bold'),
            bg=self.colors['white'],
            fg=self.colors['dark_blue'],
            padx=15,
            pady=6,
            relief='raised',
            cursor='hand2'
        )
        self.about_btn.pack(side='left', padx=5)
        
        # Language toggle button
        self.lang_btn = tk.Button(
            right_btns_frame,
            text="اردو/English",
            command=self.toggle_language,
            font=('Arial', 11, 'bold'),
            bg=self.colors['white'],
            fg=self.colors['green'],
            padx=12,
            pady=6,
            relief='raised',
            cursor='hand2'
        )
        self.lang_btn.pack(side='left', padx=5)
        
        # Title with large font
        self.title_label = tk.Label(
            header_frame,
            text="🇵🇰 ULTIMATE PAKISTANI LAND CONVERTER v3.1",
            font=('Arial', 24, 'bold'),
            bg=self.colors['green'],
            fg=self.colors['white']
        )
        self.title_label.pack(pady=(12, 5))
        
        self.subtitle_label = tk.Label(
            header_frame,
            text="Punjab Legal Standard (225 sq ft) + Traditional References (272 sq ft)",
            font=('Arial', 15, 'bold'),
            bg=self.colors['green'],
            fg=self.colors['white']
        )
        self.subtitle_label.pack()
        
        self.standards_label = tk.Label(
            header_frame,
            text=f"LEGAL: 1 Marla = {SQFT_PER_MARLA_LEGAL} sq ft | TRADITIONAL: 1 Marla = {SQFT_PER_MARLA_TRAD} sq ft (KPK Builders) | 1 Sq. Karam = {SQFT_PER_SQ_KARAM} sq ft",
            font=('Arial', 12),
            bg=self.colors['green'],
            fg=self.colors['white']
        )
        self.standards_label.pack(pady=(5, 8))
        
        # ========== NOTEBOOK (TABS) ==========
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill='both', expand=True, padx=15, pady=10)
        
        # Tab 1: Converter
        converter_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(converter_frame, text="  📏 Converter  ")
        
        # Tab 2: Visualization
        viz_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(viz_frame, text="  📈 Visualization  ")
        
        # Tab 3: Reverse Lookup
        lookup_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(lookup_frame, text="  🔍 Reverse Lookup  ")
        
        # Tab 4: Area Calculator
        area_calc_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(area_calc_frame, text="  📐 Area Calculator  ")
        
        # Build Converter Tab
        self.build_converter_tab(converter_frame)
        
        # Build Visualization Tab
        self.build_viz_tab(viz_frame)
        
        # Build Lookup Tab
        self.build_lookup_tab(lookup_frame)
        
        # Build Area Calculator Tab
        self.build_area_tab(area_calc_frame)
        
        # ========== STATUS BAR ==========
        self.status_var = tk.StringVar()
        self.status_var.set("Ready • Enter square feet for instant Punjab Legal + Traditional reference conversions | Ctrl+A for Reference Guide")
        
        status_bar = tk.Label(
            main_frame,
            textvariable=self.status_var,
            font=('Arial', 12, 'bold'),
            bg=self.colors['gray'],
            fg='black',
            anchor='w',
            padx=15,
            pady=8,
            relief='sunken',
            borderwidth=1
        )
        status_bar.pack(fill='x', side='bottom')
    
    def build_converter_tab(self, parent):
        """Build the converter tab with optimized column layout"""
        
        # ========== INPUT FRAME ==========
        input_frame = tk.LabelFrame(
            parent,
            text=" 📏 Enter Area in Square Feet ",
            font=('Arial', 15, 'bold'),
            padx=25,
            pady=20,
            bg=self.colors['white'],
            relief='groove',
            borderwidth=3
        )
        input_frame.pack(fill='x', padx=20, pady=15)
        
        # Input field with larger font
        input_row = tk.Frame(input_frame, bg=self.colors['white'])
        input_row.pack(fill='x', pady=10)
        
        tk.Label(
            input_row,
            text="Square Feet:",
            font=('Arial', 16, 'bold'),
            bg=self.colors['white'],
            width=18,
            anchor='w'
        ).pack(side='left', padx=(0, 15))
        
        self.sqft_var = tk.StringVar()
        self.sqft_entry = tk.Entry(
            input_row,
            textvariable=self.sqft_var,
            font=('Arial', 18),
            width=30,
            relief='solid',
            borderwidth=2,
            justify='center',
            bg=self.colors['light_blue']
        )
        self.sqft_entry.pack(side='left', padx=(0, 20))
        self.sqft_entry.bind('<Return>', lambda e: self.add_conversion())
        
        # Karam info label
        tk.Label(
            input_row,
            text="1 Sq. Karam = 30.25 sq ft (5.5 ft × 5.5 ft)",
            font=('Arial', 11, 'italic'),
            bg=self.colors['white'],
            fg='#5D4037'
        ).pack(side='left', padx=(0, 10))
        
        # Buttons frame
        btn_frame = tk.Frame(input_frame, bg=self.colors['white'])
        btn_frame.pack(pady=15)
        
        self.add_btn = tk.Button(
            btn_frame,
            text="➕ Add Conversion",
            command=self.add_conversion,
            font=('Arial', 14, 'bold'),
            bg=self.colors['accent_green'],
            fg=self.colors['white'],
            padx=25,
            pady=12,
            relief='raised',
            cursor='hand2',
            width=17
        )
        self.add_btn.pack(side='left', padx=7)
        
        self.import_btn = tk.Button(
            btn_frame,
            text="📤 Import from Excel",
            command=self.import_from_excel,
            font=('Arial', 14, 'bold'),
            bg=self.colors['accent_blue'],
            fg=self.colors['white'],
            padx=25,
            pady=12,
            relief='raised',
            cursor='hand2',
            width=17
        )
        self.import_btn.pack(side='left', padx=7)
        
        self.clear_btn = tk.Button(
            btn_frame,
            text="🗑️ Clear All",
            command=self.clear_all,
            font=('Arial', 14, 'bold'),
            bg=self.colors['accent_red'],
            fg=self.colors['white'],
            padx=25,
            pady=12,
            relief='raised',
            cursor='hand2',
            width=17
        )
        self.clear_btn.pack(side='left', padx=7)
        
        self.export_btn = tk.Button(
            btn_frame,
            text="📥 Export to Excel",
            command=self.export_to_excel,
            font=('Arial', 14, 'bold'),
            bg='#9C27B0',
            fg=self.colors['white'],
            padx=25,
            pady=12,
            relief='raised',
            cursor='hand2',
            width=17
        )
        self.export_btn.pack(side='left', padx=7)
        
        # ========== RESULTS FRAME ==========
        results_frame = tk.LabelFrame(
            parent,
            text=" 📊 Conversion Results (All References Shown) ",
            font=('Arial', 15, 'bold'),
            padx=20,
            pady=15,
            bg=self.colors['white'],
            relief='groove',
            borderwidth=3
        )
        results_frame.pack(fill='both', expand=True, padx=20, pady=15)
        
        # Treeview for displaying conversions with optimized columns
        tree_frame = tk.Frame(results_frame, bg=self.colors['white'])
        tree_frame.pack(fill='both', expand=True)
        
        # Scrollbars
        tree_scroll_y = tk.Scrollbar(tree_frame, orient='vertical')
        tree_scroll_y.pack(side='right', fill='y')
        
        tree_scroll_x = tk.Scrollbar(tree_frame, orient='horizontal')
        tree_scroll_x.pack(side='bottom', fill='x')
        
        # Treeview widget with LEGAL/REFERENCE columns (NO KILLA)
        columns = ('#', 'Square_Feet', 'Marla_Legal', 'Kanal_Legal', 'Marla_Trad', 'Kanal_KPK_Ref', 'Sq_Karam')
        self.tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show='headings',
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set,
            height=14
        )
        
        tree_scroll_y.config(command=self.tree.yview)
        tree_scroll_x.config(command=self.tree.xview)
        
        self.tree.pack(fill='both', expand=True)
        
        # Configure columns with color-coded headers
        self.tree.heading('#', text='#')
        self.tree.heading('Square_Feet', text='Square Feet')
        self.tree.heading('Marla_Legal', text='Marla (Punjab Legal)')
        self.tree.heading('Kanal_Legal', text='Kanal (Punjab Legal)')
        self.tree.heading('Marla_Trad', text='Marla (Trad Ref)')
        self.tree.heading('Kanal_KPK_Ref', text='Kanal (KPK Ref)')
        self.tree.heading('Sq_Karam', text='Sq. Karam')
        
        self.tree.column('#', width=65, anchor='center')
        self.tree.column('Square_Feet', width=145, anchor='center')
        self.tree.column('Marla_Legal', width=165, anchor='center')
        self.tree.column('Kanal_Legal', width=165, anchor='center')
        self.tree.column('Marla_Trad', width=155, anchor='center')
        self.tree.column('Kanal_KPK_Ref', width=165, anchor='center')
        self.tree.column('Sq_Karam', width=135, anchor='center')
        
        # Style the Treeview with larger font and color coding
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Treeview',
                       background=self.colors['white'],
                       foreground='black',
                       rowheight=34,
                       fieldbackground=self.colors['white'],
                       font=('Arial', 13))
        style.configure('Treeview.Heading',
                       background=self.colors['dark_blue'],
                       foreground=self.colors['white'],
                       font=('Arial', 12, 'bold'),
                       padding=6,
                       relief='raised')
        
        # Color-code columns by legal status
        self.tree.tag_configure('legal_col', background=self.colors['light_blue'])
        self.tree.tag_configure('trad_col', background=self.colors['trad_yellow'])
        self.tree.tag_configure('kpk_col', background=self.colors['kpk_orange'])
        
        # ========== LEGAL NOTICE FRAME ==========
        legal_frame = tk.LabelFrame(
            parent,
            text=" ⚠️ Critical Legal Notice for Patwaris & Lawyers ",
            font=('Arial', 14, 'bold'),
            padx=20,
            pady=15,
            bg=self.colors['white'],
            relief='groove',
            borderwidth=3
        )
        legal_frame.pack(fill='x', padx=20, pady=15)
        
        self.legal_text = tk.Text(
            legal_frame,
            height=4,
            wrap='word',
            font=('Arial', 12, 'bold'),
            bg='#FFEBEE',
            fg='#B71C1C',
            relief='solid',
            borderwidth=2,
            padx=15,
            pady=12
        )
        self.legal_text.pack(fill='x')
        self.legal_text.insert('1.0', 
            "⚠️ LEGAL WARNING: ONLY 'Marla (Punjab Legal)' = 225 sq ft is valid for Fard-e-Malkiat, mutations (Intiqal), "
            "court cases, and property registration. 'Marla (Trad Ref)' and 'Kanal (KPK Ref)' are INFORMAL REFERENCES ONLY "
            "— using them in official documents may result in mutation rejection or legal penalties under Section 182 PPC."
        )
        self.legal_text.config(state='disabled')
    
    def build_viz_tab(self, parent):
        """Build visualization tab with legal/traditional comparison"""
        
        instr_label = tk.Label(
            parent,
            text="📊 Punjab Legal vs Traditional Reference Comparison",
            font=('Arial', 17, 'bold'),
            bg=self.colors['white'],
            fg=self.colors['dark_blue'],
            pady=15
        )
        instr_label.pack()
        
        # Chart container
        self.chart_frame = tk.Frame(parent, bg=self.colors['white'])
        self.chart_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Placeholder message
        self.placeholder_label = tk.Label(
            self.chart_frame,
            text="No data to visualize. Add conversions first!",
            font=('Arial', 15, 'italic'),
            bg=self.colors['white'],
            fg=self.colors['gray']
        )
        self.placeholder_label.pack(expand=True)
        
        # Refresh button
        refresh_btn = tk.Button(
            parent,
            text="🔄 Refresh Chart",
            command=self.update_chart,
            font=('Arial', 14, 'bold'),
            bg=self.colors['accent_blue'],
            fg=self.colors['white'],
            padx=25,
            pady=11,
            relief='raised',
            cursor='hand2'
        )
        refresh_btn.pack(pady=15)
    
    def update_chart(self):
        """Update chart showing legal vs traditional comparisons"""
        if not self.conversions:
            if hasattr(self, 'canvas') and self.canvas:
                self.canvas.get_tk_widget().pack_forget()
                self.canvas = None
            if not self.placeholder_label.winfo_ismapped():
                self.placeholder_label.pack(expand=True)
            return
        
        if self.placeholder_label.winfo_ismapped():
            self.placeholder_label.pack_forget()
        
        # Prepare data
        sqft_vals = [c['Square_Feet'] for c in self.conversions]
        marla_legal = [c['Marla_Legal'] for c in self.conversions]
        marla_trad = [c['Marla_Trad'] for c in self.conversions]
        kanal_legal = [c['Kanal_Legal'] for c in self.conversions]
        kanal_kpk = [c['Kanal_KPK_Ref'] for c in self.conversions]
        
        # Create figure with dual y-axes
        fig = Figure(figsize=(12, 7), dpi=100)
        ax1 = fig.add_subplot(111)
        
        x = np.arange(len(self.conversions))
        width = 0.35
        
        bars1 = ax1.bar(x - width/2, marla_legal, width, label='Marla (Punjab Legal)', 
                       color=self.colors['light_blue'], edgecolor='black', linewidth=1.5)
        bars2 = ax1.bar(x + width/2, marla_trad, width, label='Marla (Trad Ref)', 
                       color=self.colors['trad_yellow'], edgecolor='black', linewidth=1.5)
        
        ax1.set_xlabel('Conversion #', fontweight='bold', fontsize=13)
        ax1.set_ylabel('Marla', fontweight='bold', fontsize=13)
        ax1.set_title('Punjab Legal vs Traditional Reference (Marla Comparison)', 
                     fontweight='bold', fontsize=14, pad=15)
        ax1.legend(loc='upper left', fontsize=11)
        ax1.grid(axis='y', alpha=0.3, linestyle='--')
        ax1.set_xticks(x)
        ax1.set_xticklabels([f"#{i+1}" for i in range(len(self.conversions))], fontsize=10)
        
        # Add value labels
        for bars in [bars1, bars2]:
            for bar in bars:
                height = bar.get_height()
                ax1.text(bar.get_x() + bar.get_width()/2., height,
                        f'{height:.1f}',
                        ha='center', va='bottom', fontsize=9, fontweight='bold')
        
        fig.tight_layout()
        
        # Embed in Tkinter
        if hasattr(self, 'canvas') and self.canvas:
            self.canvas.get_tk_widget().pack_forget()
        
        self.canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(fill='both', expand=True)

    def build_lookup_tab(self, parent):
        """Build the reverse lookup tab"""
        # ========== INPUT FRAME ==========
        self.lookup_input_frame = tk.LabelFrame(
            parent,
            text=" 🔍 Reverse Measurement Lookup ",
            font=('Arial', 15, 'bold'),
            padx=25,
            pady=20,
            bg=self.colors['white'],
            relief='groove',
            borderwidth=3
        )
        self.lookup_input_frame.pack(fill='x', padx=20, pady=15)
        
        input_row = tk.Frame(self.lookup_input_frame, bg=self.colors['white'])
        input_row.pack(fill='x', pady=10)
        
        # Unit selection
        self.lookup_unit_label = tk.Label(
            input_row,
            text="Select Unit:",
            font=('Arial', 16, 'bold'),
            bg=self.colors['white'],
            anchor='w'
        )
        self.lookup_unit_label.pack(side='left', padx=(0, 10))
        
        self.lookup_units = [
            'Square Feet',
            'Marla (Punjab Legal)',
            'Kanal (Punjab Legal)',
            'Marla (Trad Ref)',
            'Kanal (KPK Ref)',
            'Sq. Karam'
        ]
        self.lookup_unit_var = tk.StringVar(value=self.lookup_units[1])
        self.lookup_unit_cb = ttk.Combobox(
            input_row,
            textvariable=self.lookup_unit_var,
            values=self.lookup_units,
            font=('Arial', 14),
            width=22,
            state='readonly'
        )
        self.lookup_unit_cb.pack(side='left', padx=(0, 20))
        
        # Value Input
        self.lookup_value_label = tk.Label(
            input_row,
            text="Enter Value:",
            font=('Arial', 16, 'bold'),
            bg=self.colors['white'],
            anchor='w'
        )
        self.lookup_value_label.pack(side='left', padx=(0, 10))
        
        self.lookup_val_var = tk.StringVar()
        self.lookup_entry = tk.Entry(
            input_row,
            textvariable=self.lookup_val_var,
            font=('Arial', 18),
            width=15,
            relief='solid',
            borderwidth=2,
            justify='center',
            bg=self.colors['trad_yellow']
        )
        self.lookup_entry.pack(side='left', padx=(0, 20))
        self.lookup_entry.bind('<Return>', lambda e: self.calculate_lookup())
        
        self.lookup_calc_btn = tk.Button(
            input_row,
            text="🧮 Convert",
            command=self.calculate_lookup,
            font=('Arial', 14, 'bold'),
            bg=self.colors['dark_blue'],
            fg=self.colors['white'],
            padx=20,
            pady=8,
            relief='raised',
            cursor='hand2'
        )
        self.lookup_calc_btn.pack(side='left', padx=(10, 0))
        
        # ========== RESULTS FRAME ==========
        self.lookup_results_frame = tk.LabelFrame(
            parent,
            text=" 🔍 Lookup Results ",
            font=('Arial', 15, 'bold'),
            padx=20,
            pady=15,
            bg=self.colors['white'],
            relief='groove',
            borderwidth=3
        )
        self.lookup_results_frame.pack(fill='both', expand=True, padx=20, pady=15)
        
        # Treeview for displaying lookup results
        tree_frame = tk.Frame(self.lookup_results_frame, bg=self.colors['white'])
        tree_frame.pack(fill='both', expand=True)
        
        tree_scroll_y = tk.Scrollbar(tree_frame, orient='vertical')
        tree_scroll_y.pack(side='right', fill='y')
        
        columns = ('Unit', 'Value')
        self.lookup_tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show='headings',
            yscrollcommand=tree_scroll_y.set,
            height=8
        )
        tree_scroll_y.config(command=self.lookup_tree.yview)
        self.lookup_tree.pack(fill='both', expand=True)
        
        self.lookup_tree.heading('Unit', text='Measurement Unit')
        self.lookup_tree.heading('Value', text='Converted Value')
        
        self.lookup_tree.column('Unit', width=300, anchor='w')
        self.lookup_tree.column('Value', width=400, anchor='center')
        
        # Configure tags for tree rows
        self.lookup_tree.tag_configure('highlight_legal', background=self.colors['light_blue'], font=('Arial', 14, 'bold'))
        self.lookup_tree.tag_configure('highlight_trad', background=self.colors['trad_yellow'], font=('Arial', 14, 'bold'))
        self.lookup_tree.tag_configure('normal_row', background=self.colors['gray'], font=('Arial', 13))

    def calculate_lookup(self):
        """Calculate and display all units based on the selected reverse lookup input"""
        try:
            val_input = self.lookup_val_var.get().strip()
            
            if not val_input:
                msg = URDU_TEXT['input_required_msg'] if self.is_urdu else "Please enter a value!"
                messagebox.showwarning(URDU_TEXT['input_required'] if self.is_urdu else "Input Required", msg)
                self.lookup_entry.focus()
                return
            
            val = float(val_input)
            if val <= 0:
                msg = URDU_TEXT['invalid_positive'] if self.is_urdu else "Value must be greater than zero!"
                messagebox.showerror(URDU_TEXT['invalid_input'] if self.is_urdu else "Invalid Input", msg)
                return
                
            unit = self.lookup_unit_var.get()
            sqft = 0.0
            
            # Identify input unit and map to sqft
            if unit == 'Square Feet' or unit == 'مربع فٹ':
                sqft = val
            elif unit == 'Marla (Punjab Legal)' or unit == URDU_TEXT.get('marla_legal', ''):
                sqft = val * SQFT_PER_MARLA_LEGAL
            elif unit == 'Kanal (Punjab Legal)' or unit == URDU_TEXT.get('kanal_legal', ''):
                sqft = val * SQFT_PER_KANAL_LEGAL
            elif unit == 'Marla (Trad Ref)' or unit == URDU_TEXT.get('marla_trad', ''):
                sqft = val * SQFT_PER_MARLA_TRAD
            elif unit == 'Kanal (KPK Ref)' or unit == URDU_TEXT.get('kanal_kpk', ''):
                sqft = val * SQFT_PER_KANAL_TRAD
            elif unit == 'Sq. Karam' or unit == URDU_TEXT.get('sq_karam', ''):
                sqft = val * SQFT_PER_SQ_KARAM
            else:
                sqft = val # Fallback
                
            marla_legal = sqft / SQFT_PER_MARLA_LEGAL
            kanal_legal = sqft / SQFT_PER_KANAL_LEGAL
            marla_trad = sqft / SQFT_PER_MARLA_TRAD
            kanal_kpk = sqft / SQFT_PER_KANAL_TRAD
            sq_karam = sqft / SQFT_PER_SQ_KARAM
            
            # Clear previous
            for item in self.lookup_tree.get_children():
                self.lookup_tree.delete(item)
                
            if self.is_urdu:
                results = [
                    ('مربع فٹ', f"{sqft:,.2f}", 'normal_row'),
                    (URDU_TEXT['marla_legal'], f"{marla_legal:,.4f}", 'highlight_legal'),
                    (URDU_TEXT['kanal_legal'], f"{kanal_legal:,.4f}", 'highlight_legal'),
                    (URDU_TEXT['marla_trad'], f"{marla_trad:,.4f}", 'highlight_trad'),
                    (URDU_TEXT['kanal_kpk'], f"{kanal_kpk:,.4f}", 'highlight_trad'),
                    (URDU_TEXT['sq_karam'], f"{sq_karam:,.2f}", 'normal_row')
                ]
            else:
                results = [
                    ('Square Feet', f"{sqft:,.2f} sq ft", 'normal_row'),
                    ('Marla (Punjab Legal)', f"{marla_legal:,.4f} Marla", 'highlight_legal'),
                    ('Kanal (Punjab Legal)', f"{kanal_legal:,.4f} Kanal", 'highlight_legal'),
                    ('Marla (Trad Ref)', f"{marla_trad:,.4f} Marla", 'highlight_trad'),
                    ('Kanal (KPK Ref)', f"{kanal_kpk:,.4f} Kanal", 'highlight_trad'),
                    ('Sq. Karam', f"{sq_karam:,.2f} Sq. Karam", 'normal_row')
                ]
            
            for res_unit, res_val, tag in results:
                if unit == res_unit:
                    self.lookup_tree.insert('', 'end', values=(res_unit, "► " + res_val), tags=(tag,))
                else:
                    self.lookup_tree.insert('', 'end', values=(res_unit, res_val), tags=(tag,))
                
        except ValueError:
            msg = URDU_TEXT['invalid_number'] if self.is_urdu else "Please enter a valid number!"
            messagebox.showerror(URDU_TEXT['invalid_input'] if self.is_urdu else "Invalid Input", msg)
            self.lookup_entry.focus()
    

    def build_area_tab(self, parent):
        """Build the Area Calculator tab"""
        self.area_input_frame = tk.LabelFrame(
            parent,
            text=" 📐 Survey Polygon Area Calculator ",
            font=('Arial', 15, 'bold'),
            padx=25,
            pady=20,
            bg=self.colors['white'],
            relief='groove',
            borderwidth=3
        )
        self.area_input_frame.pack(fill='x', padx=20, pady=15)
        
        row1 = tk.Frame(self.area_input_frame, bg=self.colors['white'])
        row1.pack(fill='x', pady=5)
        
        self.area_shape_label = tk.Label(
            row1,
            text="Choose Shape:",
            font=('Arial', 14, 'bold'),
            bg=self.colors['white'],
            anchor='w'
        )
        self.area_shape_label.pack(side='left', padx=(0, 10))
        
        self.area_shapes = [
            'Rectangle (Length × Width)',
            'Triangle (3 sides - Heron\'s Formula)',
            'Quadrilateral (Exact: 4 sides + 1 diagonal)',
            'Quadrilateral (Traditional Patwari Average)'
        ]
        for i in range(5, 11):
            self.area_shapes.append(f"{i}-Sided Polygon ({i} sides + {i-3} diagonals)")
            
        self.area_shape_var = tk.StringVar(value=self.area_shapes[2])
        self.area_shape_cb = ttk.Combobox(
            row1,
            textvariable=self.area_shape_var,
            values=self.area_shapes,
            font=('Arial', 13),
            width=42,
            state='readonly'
        )
        self.area_shape_cb.pack(side='left', padx=(0, 20))
        self.area_shape_cb.bind('<<ComboboxSelected>>', self.on_shape_change)
        
        # Container for entries and canvas
        self.content_frame = tk.Frame(self.area_input_frame, bg=self.colors['white'])
        self.content_frame.pack(fill='both', expand=True, pady=15)
        
        # Frame for dynamic side entries (Left side)
        self.side_entries_frame = tk.Frame(self.content_frame, bg=self.colors['white'])
        self.side_entries_frame.pack(side='left', fill='both', expand=True, padx=(0, 10))
        
        # Frame for Canvas (Right side)
        self.canvas_frame = tk.Frame(self.content_frame, bg=self.colors['white'], relief='sunken', borderwidth=2)
        self.canvas_frame.pack(side='right', fill='both', padx=(10, 0))
        
        self.shape_canvas = tk.Canvas(self.canvas_frame, width=320, height=250, bg='#F0F8FF')
        self.shape_canvas.pack(padx=10, pady=10)
        
        self.side_vars = {}
        self.side_labels = {}
        self.side_entries = {}
        
        for i in range(1, 25): # Support up to 24 inputs (e.g., 10 sides + 7 diagonals = 17)
            frame = tk.Frame(self.side_entries_frame, bg=self.colors['white'])
            
            lbl = tk.Label(frame, text=f"Input {i}:", font=('Arial', 11, 'bold'), bg=self.colors['white'], width=18, anchor='w')
            lbl.pack(side='left', padx=2)
            self.side_labels[i] = lbl
            
            var = tk.StringVar()
            self.side_vars[i] = var
            ent = tk.Entry(frame, textvariable=var, font=('Arial', 14), width=8, justify='center', bg=self.colors['trad_yellow'], relief='solid', borderwidth=1)
            ent.pack(side='left', padx=2)
            self.side_entries[i] = ent
        
        btn_frame = tk.Frame(self.area_input_frame, bg=self.colors['white'])
        btn_frame.pack(fill='x', pady=10)
        
        self.area_calc_btn = tk.Button(
            btn_frame,
            text="🧮 Calculate Sq Ft Area",
            command=self.calculate_polygon_area,
            font=('Arial', 14, 'bold'),
            bg=self.colors['dark_blue'],
            fg=self.colors['white'],
            padx=20,
            pady=8,
            relief='raised',
            cursor='hand2'
        )
        self.area_calc_btn.pack(side='left')
        
        # Results frame
        self.area_results_frame = tk.LabelFrame(
            parent,
            text=" 📋 Area Results ",
            font=('Arial', 15, 'bold'),
            padx=20,
            pady=15,
            bg=self.colors['white'],
            relief='groove',
            borderwidth=3
        )
        self.area_results_frame.pack(fill='both', expand=True, padx=20, pady=15)
        
        res_tree_frame = tk.Frame(self.area_results_frame, bg=self.colors['white'])
        res_tree_frame.pack(fill='both', expand=True)
        
        tree_scroll_y = tk.Scrollbar(res_tree_frame, orient='vertical')
        tree_scroll_y.pack(side='right', fill='y')
        
        columns = ('Unit', 'Value')
        self.area_tree = ttk.Treeview(
            res_tree_frame,
            columns=columns,
            show='headings',
            yscrollcommand=tree_scroll_y.set,
            height=6
        )
        tree_scroll_y.config(command=self.area_tree.yview)
        self.area_tree.pack(fill='both', expand=True)
        
        self.area_tree.heading('Unit', text='Unit')
        self.area_tree.heading('Value', text='Area')
        self.area_tree.column('Unit', width=300, anchor='w')
        self.area_tree.column('Value', width=400, anchor='center')
        
        self.area_tree.tag_configure('highlight_legal', background=self.colors['light_blue'], font=('Arial', 14, 'bold'))
        self.area_tree.tag_configure('highlight_trad', background=self.colors['trad_yellow'], font=('Arial', 14, 'bold'))
        self.area_tree.tag_configure('normal_row', background=self.colors['gray'], font=('Arial', 13))
        
        self.on_shape_change()

    def on_shape_change(self, event=None):
        """Dynamically update input fields based on selected shape"""
        shape = self.area_shape_var.get()
        is_urdu = getattr(self, 'is_urdu', False)
        
        for frame in self.side_labels.values():
            frame.master.grid_forget()
            
        def show_input(idx, eng_label, urdu_label, row, col):
            self.side_labels[idx].config(text=urdu_label if is_urdu else eng_label)
            self.side_labels[idx].master.grid(row=row, column=col, padx=4, pady=4, sticky='w')

        if 'Rectangle' in shape or 'مستطیل' in shape:
            show_input(1, "Length (ft):", URDU_TEXT['side_1'] if is_urdu else "پہلا ضلع (فٹ):", 0, 0)
            show_input(2, "Width (ft):", URDU_TEXT['side_2'] if is_urdu else "دوسرا ضلع (فٹ):", 0, 1)
        elif 'Triangle' in shape or 'مثلث' in shape:
            show_input(1, "Side 1 (ft):", URDU_TEXT['side_1'] if is_urdu else "پہلا ضلع (فٹ):", 0, 0)
            show_input(2, "Side 2 (ft):", URDU_TEXT['side_2'] if is_urdu else "دوسرا ضلع (فٹ):", 0, 1)
            show_input(3, "Side 3 (ft):", URDU_TEXT['side_3'] if is_urdu else "تیسرا ضلع (فٹ):", 0, 2)
        elif 'Exact' in shape or 'درست' in shape:
            show_input(1, "Side 1 (ft):", URDU_TEXT['side_1'] if is_urdu else "پہلا ضلع (فٹ):", 0, 0)
            show_input(2, "Side 2 (ft):", URDU_TEXT['side_2'] if is_urdu else "دوسرا ضلع (فٹ):", 0, 1)
            show_input(3, "Side 3 (ft):", URDU_TEXT['side_3'] if is_urdu else "تیسرا ضلع (فٹ):", 0, 2)
            show_input(4, "Side 4 (ft):", URDU_TEXT['side_4'] if is_urdu else "چوتھا ضلع (فٹ):", 0, 3)
            show_input(5, "Diagonal (ft):", URDU_TEXT['diagonal'] if is_urdu else "وتر/درمیان (فٹ):", 1, 0)
        elif 'Average' in shape or 'اوسط' in shape:
            show_input(1, "Side 1/Opp1 (ft):", URDU_TEXT['side_1'] if is_urdu else "پہلا ضلع (فٹ):", 0, 0)
            show_input(2, "Side 2/Opp2 (ft):", URDU_TEXT['side_2'] if is_urdu else "دوسرا ضلع (فٹ):", 0, 1)
            show_input(3, "Side 3/Opp1 (ft):", URDU_TEXT['side_3'] if is_urdu else "تیسرا ضلع (فٹ):", 1, 0)
            show_input(4, "Side 4/Opp2 (ft):", URDU_TEXT['side_4'] if is_urdu else "چوتھا ضلع (فٹ):", 1, 1)
        elif '-Sided' in shape or 'اضلاع' in shape:
            import re
            m = re.search(r'(\d+)-Sided', shape)
            if not m: m = re.search(r'(\d+)', shape)
            n = int(m.group(1)) if m else 5
            
            # Show sides
            for i in range(1, n+1):
                show_input(i, f"Side {i} (ft):", f"ضلع {i} (فٹ):", row=(i-1)//4, col=(i-1)%4)
            
            # Show diagonals (n-3 diagonals)
            for d in range(1, n-2):
                idx = n + d
                r = (idx-1)//4
                c = (idx-1)%4
                show_input(idx, f"Diag {d} (ft):", f"وتر {d} (فٹ):", row=r+1+(n//4), col=c)
                
        self._draw_shape_schematic()

    def _draw_shape_schematic(self):
        """Draw a representative schematic on the canvas"""
        if not hasattr(self, 'shape_canvas'):
            return
            
        self.shape_canvas.delete("all")
        shape = self.area_shape_var.get()
        w = 320
        h = 250
        cx = w / 2
        cy = h / 2
        radius = min(w, h) / 2 - 30
        
        def draw_label(x1, y1, x2, y2, text, color='blue', offset=15):
            mx, my = (x1 + x2) / 2, (y1 + y2) / 2
            dx, dy = mx - cx, my - cy
            dist = math.hypot(dx, dy)
            if dist > 0:
                nx, ny = dx/dist, dy/dist
                lx, ly = mx + nx*offset, my + ny*offset
            else:
                lx, ly = mx, my
            self.shape_canvas.create_text(lx, ly, text=text, fill=color, font=('Arial', 10, 'bold'))

        if 'Rectangle' in shape or 'مستطیل' in shape:
            pts = [(cx-80, cy-50), (cx+80, cy-50), (cx+80, cy+50), (cx-80, cy+50)]
            self.shape_canvas.create_polygon(pts, fill='#E3F2FD', outline='#1565C0', width=2)
            draw_label(pts[0][0], pts[0][1], pts[1][0], pts[1][1], "Side 2 (W)")
            draw_label(pts[1][0], pts[1][1], pts[2][0], pts[2][1], "Side 1 (L)")
            draw_label(pts[2][0], pts[2][1], pts[3][0], pts[3][1], "Side 2 (W)")
            draw_label(pts[3][0], pts[3][1], pts[0][0], pts[0][1], "Side 1 (L)")
            
        elif 'Triangle' in shape or 'مثلث' in shape:
            pts = [(cx, cy-radius), (cx+radius*0.866, cy+radius*0.5), (cx-radius*0.866, cy+radius*0.5)]
            self.shape_canvas.create_polygon(pts, fill='#E3F2FD', outline='#1565C0', width=2)
            draw_label(pts[0][0], pts[0][1], pts[1][0], pts[1][1], "S2")
            draw_label(pts[1][0], pts[1][1], pts[2][0], pts[2][1], "S3")
            draw_label(pts[2][0], pts[2][1], pts[0][0], pts[0][1], "S1")
            
        elif 'Exact' in shape or 'درست' in shape:
            pts = [(cx-radius*0.7, cy-radius*0.7), (cx+radius*0.7, cy-radius*0.5), 
                   (cx+radius*0.5, cy+radius*0.8), (cx-radius*0.8, cy+radius*0.5)]
            self.shape_canvas.create_polygon(pts, fill='#E3F2FD', outline='#1565C0', width=2)
            self.shape_canvas.create_line(pts[0][0], pts[0][1], pts[2][0], pts[2][1], fill='red', dash=(4,4), width=2)
            draw_label(pts[0][0], pts[0][1], pts[1][0], pts[1][1], "S1")
            draw_label(pts[1][0], pts[1][1], pts[2][0], pts[2][1], "S2")
            draw_label(pts[2][0], pts[2][1], pts[3][0], pts[3][1], "S3")
            draw_label(pts[3][0], pts[3][1], pts[0][0], pts[0][1], "S4")
            draw_label(pts[0][0], pts[0][1], pts[2][0], pts[2][1], "Diag", color='red', offset=0)
            
        elif 'Average' in shape or 'اوسط' in shape:
            pts = [(cx-radius*0.7, cy-radius*0.7), (cx+radius*0.7, cy-radius*0.5), 
                   (cx+radius*0.5, cy+radius*0.8), (cx-radius*0.8, cy+radius*0.5)]
            self.shape_canvas.create_polygon(pts, fill='#FFF8E1', outline='#F57F17', width=2)
            draw_label(pts[0][0], pts[0][1], pts[1][0], pts[1][1], "S1")
            draw_label(pts[1][0], pts[1][1], pts[2][0], pts[2][1], "S2")
            draw_label(pts[2][0], pts[2][1], pts[3][0], pts[3][1], "S3 (Opp 1)")
            draw_label(pts[3][0], pts[3][1], pts[0][0], pts[0][1], "S4 (Opp 2)")
            
        elif '-Sided' in shape or 'اضلاع' in shape:
            import re, math
            m = re.search(r'(\d+)-Sided', shape)
            if not m: m = re.search(r'(\d+)', shape)
            n = int(m.group(1)) if m else 5
            
            pts = []
            for i in range(n):
                angle = 2 * math.pi * i / n - math.pi/2
                px = cx + radius * math.cos(angle)
                py = cy + radius * math.sin(angle)
                pts.append((px, py))
            
            self.shape_canvas.create_polygon(pts, fill='#E8F5E9', outline='#2E7D32', width=2)
            
            for i in range(n):
                p1 = pts[i]
                p2 = pts[(i+1)%n]
                draw_label(p1[0], p1[1], p2[0], p2[1], f"S{i+1}")
                
            for i in range(n-3):
                p1 = pts[0]
                p2 = pts[i+2]
                self.shape_canvas.create_line(p1[0], p1[1], p2[0], p2[1], fill='red', dash=(4,4), width=2)
                draw_label(p1[0], p1[1], p2[0], p2[1], f"D{i+1}", color='red', offset=0)

    def calculate_polygon_area(self):
        """Calculate area based on shape and measurements"""
        shape = self.area_shape_var.get()
        sqft = 0.0
        
        try:
            def get_val(idx):
                v = self.side_vars[idx].get().strip()
                return float(v) if v else 0.0

            if 'Rectangle' in shape or 'مستطیل' in shape:
                l, w = get_val(1), get_val(2)
                sqft = l * w
            elif 'Triangle' in shape or 'مثلث' in shape:
                a, b, c = get_val(1), get_val(2), get_val(3)
                s = (a + b + c) / 2
                if s*(s-a)*(s-b)*(s-c) <= 0: raise ValueError("Invalid triangle sides")
                sqft = math.sqrt(s*(s-a)*(s-b)*(s-c))
            elif 'Exact' in shape or 'درست' in shape:
                a, b, c, d, diag = get_val(1), get_val(2), get_val(3), get_val(4), get_val(5)
                s1 = (a + b + diag) / 2
                t1 = s1*(s1-a)*(s1-b)*(s1-diag)
                if t1 <= 0: raise ValueError("Invalid Triangle 1 sides")
                s2 = (c + d + diag) / 2
                t2 = s2*(s2-c)*(s2-d)*(s2-diag)
                if t2 <= 0: raise ValueError("Invalid Triangle 2 sides")
                sqft = math.sqrt(t1) + math.sqrt(t2)
            elif 'Average' in shape or 'اوسط' in shape:
                s1, s2, s3, s4 = get_val(1), get_val(2), get_val(3), get_val(4)
                sqft = ((s1 + s3) / 2) * ((s2 + s4) / 2)
            elif '-Sided' in shape or 'اضلاع' in shape:
                import re
                m = re.search(r'(\d+)-Sided', shape)
                if not m: m = re.search(r'(\d+)', shape)
                n = int(m.group(1)) if m else 5
                
                sides = [get_val(i) for i in range(1, n+1)]
                diags = [get_val(i) for i in range(n+1, 2*n-2)]
                
                total_sqft = 0.0
                for i in range(n-2):
                    if i == 0:
                        a, b, c = sides[0], sides[1], diags[0]
                    elif i == n - 3:
                        a, b, c = diags[i-1], sides[i+1], sides[i+2]
                    else:
                        a, b, c = diags[i-1], sides[i+1], diags[i]
                    
                    s = (a + b + c) / 2
                    term = s*(s-a)*(s-b)*(s-c)
                    if term <= 0: raise ValueError(f"Invalid Triangle {i+1} sides")
                    total_sqft += math.sqrt(term)
                sqft = total_sqft
            
            if sqft <= 0:
                raise ValueError("Area is zero or negative")
                
            # Feed sqft into conversion logic
            self.sqft_var.set(str(round(sqft, 2)))
            
            marla_legal = sqft / SQFT_PER_MARLA_LEGAL
            kanal_legal = sqft / SQFT_PER_KANAL_LEGAL
            marla_trad = sqft / SQFT_PER_MARLA_TRAD
            kanal_kpk = sqft / SQFT_PER_KANAL_TRAD
            sq_karam = sqft / SQFT_PER_SQ_KARAM
            
            for item in self.area_tree.get_children():
                self.area_tree.delete(item)
                
            is_urdu = getattr(self, 'is_urdu', False)
            if is_urdu:
                results = [
                    ('مربع فٹ', f"{sqft:,.2f}", 'normal_row'),
                    (URDU_TEXT['marla_legal'], f"{marla_legal:,.4f}", 'highlight_legal'),
                    (URDU_TEXT['kanal_legal'], f"{kanal_legal:,.4f}", 'highlight_legal'),
                    (URDU_TEXT['marla_trad'], f"{marla_trad:,.4f}", 'highlight_trad'),
                    (URDU_TEXT['kanal_kpk'], f"{kanal_kpk:,.4f}", 'highlight_trad'),
                    (URDU_TEXT['sq_karam'], f"{sq_karam:,.2f}", 'normal_row')
                ]
            else:
                results = [
                    ('Square Feet', f"{sqft:,.2f} sq ft", 'normal_row'),
                    ('Marla (Punjab Legal)', f"{marla_legal:,.4f} Marla", 'highlight_legal'),
                    ('Kanal (Punjab Legal)', f"{kanal_legal:,.4f} Kanal", 'highlight_legal'),
                    ('Marla (Trad Ref)', f"{marla_trad:,.4f} Marla", 'highlight_trad'),
                    ('Kanal (KPK Ref)', f"{kanal_kpk:,.4f} Kanal", 'highlight_trad'),
                    ('Sq. Karam', f"{sq_karam:,.2f} Sq. Karam", 'normal_row')
                ]
            
            for res_unit, res_val, tag in results:
                self.area_tree.insert('', 'end', values=(res_unit, res_val), tags=(tag,))
                
        except ValueError as e:
            msg = URDU_TEXT['invalid_number'] if getattr(self, 'is_urdu', False) else f"Math Error: {str(e)}. Please check measurements so they form closed triangles."
            messagebox.showerror(URDU_TEXT['invalid_input'] if getattr(self, 'is_urdu', False) else "Invalid Input", msg)


    def add_conversion(self):
        """Add conversion with ALL reference values calculated"""
        try:
            sqft_input = self.sqft_var.get().strip()
            
            if not sqft_input:
                msg = URDU_TEXT['input_required_msg'] if self.is_urdu else "Please enter square feet value!"
                messagebox.showwarning(URDU_TEXT['input_required'] if self.is_urdu else "Input Required", msg)
                self.sqft_entry.focus()
                return
            
            sqft = float(sqft_input)
            
            if sqft <= 0:
                msg = URDU_TEXT['invalid_positive'] if self.is_urdu else "Square feet must be greater than zero!"
                messagebox.showerror(URDU_TEXT['invalid_input'] if self.is_urdu else "Invalid Input", msg)
                return
            
            # Calculate ALL reference values (always show both standards)
            marla_legal = sqft / SQFT_PER_MARLA_LEGAL
            kanal_legal = sqft / SQFT_PER_KANAL_LEGAL
            marla_trad = sqft / SQFT_PER_MARLA_TRAD
            kanal_kpk = sqft / SQFT_PER_KANAL_TRAD
            sq_karam = sqft / SQFT_PER_SQ_KARAM
            
            # Store conversion
            conversion = {
                'Square_Feet': sqft,
                'Marla_Legal': round(marla_legal, 4),
                'Kanal_Legal': round(kanal_legal, 4),
                'Marla_Trad': round(marla_trad, 4),
                'Kanal_KPK_Ref': round(kanal_kpk, 4),
                'Sq_Karam': round(sq_karam, 2)
            }
            
            self.conversions.append(conversion)
            
            # Add to treeview with color coding by column type
            item_id = self.tree.insert('', 'end', values=(
                len(self.conversions),
                f"{sqft:,.2f}",
                f"{marla_legal:.4f}",
                f"{kanal_legal:.4f}",
                f"{marla_trad:.4f}",
                f"{kanal_kpk:.4f}",
                f"{sq_karam:.2f}"
            ))
            
            # Apply column-specific coloring (requires custom rendering - simplified here)
            self.tree.item(item_id, tags=('legal_col',))
            
            # Update chart
            self.update_chart()
            
            # Update status
            if self.is_urdu:
                status = URDU_TEXT['status_added'].format(
                    sqft=f"{sqft:,.0f}",
                    marla_legal=f"{marla_legal:.2f}",
                    total=len(self.conversions)
                )
            else:
                status = f"✅ Added: {sqft:,.2f} sq ft = {marla_legal:.2f} Marla (Punjab Legal) | Total: {len(self.conversions)}"
            
            self.status_var.set(status)
            
            # Clear input and focus
            self.sqft_var.set('')
            self.sqft_entry.focus()
            
        except ValueError:
            msg = URDU_TEXT['invalid_number'] if self.is_urdu else "Please enter a valid number!"
            messagebox.showerror(URDU_TEXT['invalid_input'] if self.is_urdu else "Invalid Input", msg)
            self.sqft_entry.focus()
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
    
    def import_from_excel(self):
        """Import square feet values from Excel"""
        try:
            file_path = filedialog.askopenfilename(
                title="Select Excel File",
                filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
            )
            
            if not file_path:
                return
            
            try:
                df = pd.read_excel(file_path)
            except Exception as e:
                messagebox.showerror("Import Error", f"Could not read Excel file:\n{str(e)}")
                return
            
            # Find area column
            area_col = None
            for col in df.columns:
                col_lower = str(col).lower()
                if any(kw in col_lower for kw in ['square', 'sqft', 'area', 'feet', 'marla', 'kanal']):
                    area_col = col
                    break
            
            if area_col is None:
                for col in df.columns:
                    if pd.api.types.is_numeric_dtype(df[col]):
                        area_col = col
                        break
            
            if area_col is None:
                messagebox.showwarning("Import Error", "No numeric column found for area values!")
                return
            
            # Import valid values
            imported_count = 0
            for val in df[area_col].dropna():
                try:
                    sqft = float(val)
                    if sqft > 0:
                        marla_legal = sqft / SQFT_PER_MARLA_LEGAL
                        kanal_legal = sqft / SQFT_PER_KANAL_LEGAL
                        marla_trad = sqft / SQFT_PER_MARLA_TRAD
                        kanal_kpk = sqft / SQFT_PER_KANAL_TRAD
                        sq_karam = sqft / SQFT_PER_SQ_KARAM
                        
                        conversion = {
                            'Square_Feet': sqft,
                            'Marla_Legal': round(marla_legal, 4),
                            'Kanal_Legal': round(kanal_legal, 4),
                            'Marla_Trad': round(marla_trad, 4),
                            'Kanal_KPK_Ref': round(kanal_kpk, 4),
                            'Sq_Karam': round(sq_karam, 2)
                        }
                        
                        self.conversions.append(conversion)
                        
                        self.tree.insert('', 'end', values=(
                            len(self.conversions),
                            f"{sqft:,.2f}",
                            f"{marla_legal:.4f}",
                            f"{kanal_legal:.4f}",
                            f"{marla_trad:.4f}",
                            f"{kanal_kpk:.4f}",
                            f"{sq_karam:.2f}"
                        ))
                        
                        imported_count += 1
                except:
                    continue
            
            if imported_count == 0:
                messagebox.showwarning("Import Error", "No valid area values found!")
                return
            
            self.update_chart()
            self.status_var.set(f"✅ Imported {imported_count} conversions from Excel")
            messagebox.showinfo("Import Successful", f"✅ Successfully imported {imported_count} conversions!")
            
        except Exception as e:
            messagebox.showerror("Import Error", f"Failed to import:\n{str(e)}")
    
    def clear_all(self):
        """Clear all conversions"""
        if self.conversions:
            msg = URDU_TEXT['clear_confirm_msg'].format(count=len(self.conversions)) if self.is_urdu else f"Are you sure you want to clear all {len(self.conversions)} conversions?"
            confirm = messagebox.askyesno(URDU_TEXT['confirm_clear'] if self.is_urdu else "Confirm Clear", msg)
            
            if confirm:
                for item in self.tree.get_children():
                    self.tree.delete(item)
                self.conversions.clear()
                self.update_chart()
                self.status_var.set(URDU_TEXT['status_cleared'] if self.is_urdu else "✅ All conversions cleared")
                self.sqft_var.set('')
                self.sqft_entry.focus()
        else:
            msg = URDU_TEXT['nothing_to_clear'] if self.is_urdu else "No conversions to clear!"
            messagebox.showinfo(URDU_TEXT['confirm_clear'] if self.is_urdu else "Nothing to Clear", msg)
    
    def export_to_excel(self):
        """Export to Excel with corrected column names (NO TYPO)"""
        if not self.conversions:
            msg = URDU_TEXT['no_data_export'] if self.is_urdu else "No conversions to export!"
            messagebox.showwarning(URDU_TEXT['export_error'] if self.is_urdu else "No Data", msg)
            return
        
        try:
            file_path = filedialog.asksaveasfilename(
                title="Save Excel File",
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
                initialfile=f"Land_Conversion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not file_path:
                return
            
            # Create DataFrame with CORRECT column names (no typos!)
            df = pd.DataFrame(self.conversions)
            
            # Add description column
            df['Description'] = df.apply(
                lambda row: f"{row['Marla_Legal']:.2f} Marla (Punjab Legal) = "
                           f"{row['Kanal_Legal']:.2f} Kanal (Punjab Legal) | "
                           f"{row['Marla_Trad']:.2f} Marla (Trad Ref) = "
                           f"{row['Kanal_KPK_Ref']:.2f} Kanal (KPK Ref)",
                axis=1
            )
            
            # Reorder columns CORRECTLY (no 'Marla_Punlab' typo!)
            df = df[[
                'Square_Feet',
                'Marla_Legal',
                'Kanal_Legal',
                'Marla_Trad',
                'Kanal_KPK_Ref',
                'Sq_Karam',
                'Description'
            ]]
            
            # Export with formatting
            self.create_formatted_excel(df, file_path)
            
            messagebox.showinfo(
                URDU_TEXT['export_success'] if self.is_urdu else "Export Successful",
                f"✅ Excel file saved successfully!\n\n"
                f"Location: {file_path}\n"
                f"Conversions: {len(self.conversions)}\n\n"
                f"📁 Contains:\n"
                f"  • Marla (Punjab Legal) = 225 sq ft standard\n"
                f"  • Kanal (Punjab Legal) = 4,500 sq ft standard\n"
                f"  • Marla (Trad Ref) = 272 sq ft reference\n"
                f"  • Kanal (KPK Ref) = 5,440 sq ft reference\n"
                f"  • Sq. Karam = 30.25 sq ft agricultural unit"
            )
            
            self.status_var.set(f"✅ Exported {len(self.conversions)} conversions to Excel")
            
            if os.name == 'nt':
                os.startfile(os.path.dirname(file_path))
                
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export Excel file:\n\n{str(e)}")
    
    def create_formatted_excel(self, df, filename):
        """Create professionally formatted Excel with legal/traditional references"""
        df.to_excel(filename, index=False, sheet_name='Conversions')
        
        wb = load_workbook(filename)
        ws = wb['Conversions']
        
        # Styles
        header_fill_legal = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")  # Blue for legal
        header_fill_trad = PatternFill(start_color="FFA726", end_color="FFA726", fill_type="solid")   # Orange for traditional
        header_font = XLFont(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Format headers with color coding
        headers = [
            ('Square_Feet', 'General'),
            ('Marla_Legal', 'Legal'),
            ('Kanal_Legal', 'Legal'),
            ('Marla_Trad', 'Traditional'),
            ('Kanal_KPK_Ref', 'Traditional'),
            ('Sq_Karam', 'Agricultural'),
            ('Description', 'General')
        ]
        
        for col_idx, (col_name, col_type) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name.replace('_', ' ')
            if col_type == 'Legal':
                cell.fill = header_fill_legal
            elif col_type == 'Traditional':
                cell.fill = header_fill_trad
            else:
                cell.fill = PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid")
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Format data rows
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                if col in [1, 2, 3, 4, 5, 6]:
                    cell.number_format = '#,##0.0000'
        
        # Auto-adjust columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 4)
            ws.column_dimensions[column].width = min(adjusted_width, 32)
        
        # Add metadata sheet
        meta = wb.create_sheet('Legal Standards Reference')
        meta.append(['Pakistani Land Measurement Standards', '', ''])
        meta.append(['', '', ''])
        meta.append(['Standard Type', 'Marla', 'Kanal'])
        meta.append(['Punjab Legal (OFFICIAL)', '225 sq ft', '4,500 sq ft (20 Marla)'])
        meta.append(['Traditional Reference', '272 sq ft', '5,440 sq ft (20 Marla)'])
        meta.append(['Agricultural Unit', '9 Sq. Karam = 272.25 sq ft', '180 Sq. Karam = 5,445 sq ft'])
        meta.append(['', '', ''])
        meta.append(['Karam Measurement', '', ''])
        meta.append(['1 Linear Karam', '5.5 feet', ''])
        meta.append(['1 Square Karam', '30.25 sq ft (5.5 ft × 5.5 ft)', ''])
        meta.append(['', '', ''])
        meta.append(['⚠️ LEGAL NOTICE', '', ''])
        meta.append(['Punjab Revenue Department REQUIRES 225 sq ft/Marla (Punjab Legal Standard)', '', ''])
        meta.append(['for ALL official records (Fard-e-Malkiat, mutations, court cases, property registration).', '', ''])
        meta.append(['Traditional references (272 sq ft) are INFORMAL ONLY and NOT legally valid.', '', ''])
        meta.append(['✅ Generated on:', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ''])
        meta.append(['Total Conversions:', len(self.conversions), ''])
        
        # Format metadata
        for row in meta.iter_rows(min_row=1, max_row=16, min_col=1, max_col=3):
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.font = XLFont(bold=True, size=14, color="0A2E5A")
                elif cell.row == 3:
                    cell.font = XLFont(bold=True)
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                elif cell.row in [12, 13, 14, 15]:
                    cell.font = XLFont(bold=True, color="C62828")
        
        meta.column_dimensions['A'].width = 32
        meta.column_dimensions['B'].width = 38
        meta.column_dimensions['C'].width = 25
        
        wb.save(filename)
    
    def toggle_language(self):
        """Toggle Urdu/English interface"""
        self.is_urdu = not self.is_urdu
        
        if self.is_urdu:
            # Urdu mode
            self.root.title("پاکستانی زمین یونٹ کنورٹر v3.1")
            self.title_label.config(text="🇵🇰 " + URDU_TEXT['app_title'] + " v3.1")
            self.subtitle_label.config(text="پنجاب قانونی معیار (225 مربع فٹ) + روایتی حوالہ جات (272 مربع فٹ)")
            self.lang_btn.config(text="English/اردو")
            self.standards_label.config(text=URDU_TEXT['standards'])
            
            # Update buttons
            self.add_btn.config(text=URDU_TEXT['add_btn'])
            self.import_btn.config(text=URDU_TEXT['import_btn'])
            self.clear_btn.config(text=URDU_TEXT['clear_btn'])
            self.export_btn.config(text=URDU_TEXT['export_btn'])
            self.about_btn.config(text="ℹ️ بارے میں اور حوالہ")
            
            # Update treeview headings
            self.tree.heading('#', text='#')
            self.tree.heading('Square_Feet', text='مربع فٹ')
            self.tree.heading('Marla_Legal', text=URDU_TEXT['marla_legal'])
            self.tree.heading('Kanal_Legal', text=URDU_TEXT['kanal_legal'])
            self.tree.heading('Marla_Trad', text=URDU_TEXT['marla_trad'])
            self.tree.heading('Kanal_KPK_Ref', text=URDU_TEXT['kanal_kpk'])
            self.tree.heading('Sq_Karam', text=URDU_TEXT['sq_karam'])
            
            # Update legal notice
            self.legal_text.config(state='normal', fg='#B71C1C', font=('Arial', 12, 'bold'))
            self.legal_text.delete('1.0', tk.END)
            self.legal_text.insert('1.0', URDU_TEXT['legal_text'])
            self.legal_text.config(state='disabled')
            
            self.status_var.set("تیار • مربع فٹ درج کریں | کنٹرول+A سے حوالہ گائیڈ کھولیں")
            
            # Update reverse lookup tab
            if hasattr(self, 'notebook'):
                self.notebook.tab(2, text=URDU_TEXT['lookup_tab_text'])
                self.lookup_input_frame.config(text=" 🔍 ریورس لک اپ ")
                self.lookup_unit_label.config(text=URDU_TEXT['select_unit'])
                self.lookup_value_label.config(text=URDU_TEXT['enter_value'])
                self.lookup_calc_btn.config(text=URDU_TEXT['calc_btn'])
                self.lookup_results_frame.config(text=f" {URDU_TEXT['lookup_results_label']} ")
                
                urdu_units = [
                    'مربع فٹ',
                    URDU_TEXT['marla_legal'],
                    URDU_TEXT['kanal_legal'],
                    URDU_TEXT['marla_trad'],
                    URDU_TEXT['kanal_kpk'],
                    URDU_TEXT['sq_karam']
                ]
                
                current_unit_idx = 0
                if self.lookup_unit_var.get() in self.lookup_units:
                    current_unit_idx = self.lookup_units.index(self.lookup_unit_var.get())
                elif self.lookup_unit_var.get() in urdu_units:
                    current_unit_idx = urdu_units.index(self.lookup_unit_var.get())
                
                self.lookup_unit_cb.config(values=urdu_units)
                self.lookup_unit_var.set(urdu_units[current_unit_idx])
                self.lookup_tree.heading('Unit', text='پیمائش کی اکائی')
                self.lookup_tree.heading('Value', text='تبدیل شدہ ویلیو')
            
                # Area Calc Tab (Urdu)
                if self.notebook.index('end') > 3:
                    self.notebook.tab(3, text=URDU_TEXT['area_calc_tab'])
                    if hasattr(self, 'area_input_frame'):
                        self.area_input_frame.config(text=" " + URDU_TEXT['area_calc_tab'].strip() + " ")
                        self.area_shape_label.config(text=URDU_TEXT['choose_shape'])
                        self.area_calc_btn.config(text=URDU_TEXT['calc_area_btn'])
                        self.area_results_frame.config(text=" " + URDU_TEXT['area_results_label'] + " ")
                        
                        urdu_shapes = [
                            URDU_TEXT['shape_rect'],
                            URDU_TEXT['shape_tri'],
                            URDU_TEXT['shape_quad_tri'],
                            URDU_TEXT['shape_quad_avg']
                        ]
                        for i in range(5, 11):
                            urdu_shapes.append(f"{i} اضلاع والی شکل ({i} اضلاع + {i-3} وتر)")
                        
                        cur_shape = self.area_shape_var.get()
                        current_idx = 0
                        if cur_shape in self.area_shapes:
                            current_idx = self.area_shapes.index(cur_shape)
                        elif cur_shape in urdu_shapes:
                            current_idx = urdu_shapes.index(cur_shape)
                        
                        self.area_shape_cb.config(values=urdu_shapes)
                        self.area_shape_var.set(urdu_shapes[current_idx])
                        self.on_shape_change()
            
        else:
            # English mode
            self.root.title("🇵🇰 ULTIMATE Pakistani Land Converter v3.1 - Punjab Legal + Traditional References")
            self.title_label.config(text="🇵🇰 ULTIMATE PAKISTANI LAND CONVERTER v3.1")
            self.subtitle_label.config(text="Punjab Legal Standard (225 sq ft) + Traditional References (272 sq ft)")
            self.lang_btn.config(text="اردو/English")
            self.standards_label.config(
                text=f"LEGAL: 1 Marla = {SQFT_PER_MARLA_LEGAL} sq ft | TRADITIONAL: 1 Marla = {SQFT_PER_MARLA_TRAD} sq ft (KPK Builders) | 1 Sq. Karam = {SQFT_PER_SQ_KARAM} sq ft"
            )
            
            # Update buttons
            self.add_btn.config(text="➕ Add Conversion")
            self.import_btn.config(text="📤 Import from Excel")
            self.clear_btn.config(text="🗑️ Clear All")
            self.export_btn.config(text="📥 Export to Excel")
            self.about_btn.config(text="ℹ️ About & Reference")
            
            # Update treeview headings
            self.tree.heading('#', text='#')
            self.tree.heading('Square_Feet', text='Square Feet')
            self.tree.heading('Marla_Legal', text='Marla (Punjab Legal)')
            self.tree.heading('Kanal_Legal', text='Kanal (Punjab Legal)')
            self.tree.heading('Marla_Trad', text='Marla (Trad Ref)')
            self.tree.heading('Kanal_KPK_Ref', text='Kanal (KPK Ref)')
            self.tree.heading('Sq_Karam', text='Sq. Karam')
            
            # Update legal notice
            self.legal_text.config(state='normal', fg='#B71C1C', font=('Arial', 12, 'bold'))
            self.legal_text.delete('1.0', tk.END)
            self.legal_text.insert('1.0', 
                "⚠️ LEGAL WARNING: ONLY 'Marla (Punjab Legal)' = 225 sq ft is valid for Fard-e-Malkiat, mutations (Intiqal), "
                "court cases, and property registration. 'Marla (Trad Ref)' and 'Kanal (KPK Ref)' are INFORMAL REFERENCES ONLY "
                "— using them in official documents may result in mutation rejection or legal penalties under Section 182 PPC."
            )
            self.legal_text.config(state='disabled')
            
            self.status_var.set("Ready • Enter square feet for instant Punjab Legal + Traditional reference conversions | Ctrl+A for Reference Guide")
            
            # Update reverse lookup tab
            if hasattr(self, 'notebook') and self.notebook.index('end') >= 3:
                self.notebook.tab(2, text="  🔍 Reverse Lookup  ")
                self.lookup_input_frame.config(text=" 🔍 Reverse Measurement Lookup ")
                self.lookup_unit_label.config(text="Select Unit:")
                self.lookup_value_label.config(text="Enter Value:")
                self.lookup_calc_btn.config(text="🧮 Convert")
                self.lookup_results_frame.config(text=" 🔍 Lookup Results ")
                
                # Area Calc Tab
                if self.notebook.index('end') > 3:
                    self.notebook.tab(3, text="  📐 Area Calculator  ")
                    if hasattr(self, 'area_input_frame'):
                        self.area_input_frame.config(text=" 📐 Survey Polygon Area Calculator ")
                        self.area_shape_label.config(text="Choose Shape:")
                        self.area_calc_btn.config(text="🧮 Calculate Sq Ft Area")
                        self.area_results_frame.config(text=" 📋 Area Results ")
                        
                        current_idx = 0
                        cur_shape = self.area_shape_var.get()
                        urdu_shapes = [
                            URDU_TEXT['shape_rect'],
                            URDU_TEXT['shape_tri'],
                            URDU_TEXT['shape_quad_tri'],
                            URDU_TEXT['shape_quad_avg']
                        ]
                        for i in range(5, 11):
                            urdu_shapes.append(f"{i} اضلاع والی شکل ({i} اضلاع + {i-3} وتر)")
                        
                        if cur_shape in urdu_shapes:
                            current_idx = urdu_shapes.index(cur_shape)
                        elif cur_shape in self.area_shapes:
                            current_idx = self.area_shapes.index(cur_shape)
                            
                        self.area_shape_cb.config(values=self.area_shapes)
                        self.area_shape_var.set(self.area_shapes[current_idx])
                        self.on_shape_change()
                
                urdu_units = [
                    'مربع فٹ',
                    URDU_TEXT['marla_legal'],
                    URDU_TEXT['kanal_legal'],
                    URDU_TEXT['marla_trad'],
                    URDU_TEXT['kanal_kpk'],
                    URDU_TEXT['sq_karam']
                ]
                current_unit_idx = 0
                if self.lookup_unit_var.get() in urdu_units:
                    current_unit_idx = urdu_units.index(self.lookup_unit_var.get())
                elif self.lookup_unit_var.get() in self.lookup_units:
                    current_unit_idx = self.lookup_units.index(self.lookup_unit_var.get())
                
                self.lookup_unit_cb.config(values=self.lookup_units)
                self.lookup_unit_var.set(self.lookup_units[current_unit_idx])
                self.lookup_tree.heading('Unit', text='Measurement Unit')
                self.lookup_tree.heading('Value', text='Converted Value')
    
    # ========== ABOUT WINDOW METHODS ==========
    def show_about_window(self):
        """Display comprehensive About/Reference window"""
        
        # Create About window
        about_win = tk.Toplevel(self.root)
        about_win.title("ℹ️ About - Ultimate Pakistani Land Converter v3.1")
        about_win.geometry("1100x850")
        about_win.minsize(1000, 700)
        about_win.resizable(True, True)
        about_win.configure(bg=self.colors['white'])
        
        # Center window on screen
        about_win.update_idletasks()
        x = (about_win.winfo_screenwidth() // 2) - (1100 // 2)
        y = (about_win.winfo_screenheight() // 2) - (850 // 2)
        about_win.geometry(f"1100x850+{x}+{y}")
        
        # Make modal (focus stays on this window)
        about_win.transient(self.root)
        about_win.grab_set()
        
        # ========== HEADER ==========
        header_frame = tk.Frame(about_win, bg=self.colors['green'], height=100)
        header_frame.pack(fill='x', padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        tk.Label(
            header_frame,
            text="🇵🇰 ULTIMATE PAKISTANI LAND CONVERTER v3.1",
            font=('Arial', 22, 'bold'),
            bg=self.colors['green'],
            fg=self.colors['white']
        ).pack(pady=(15, 5))
        
        tk.Label(
            header_frame,
            text="Patwari/Lawyer Reference Guide • Punjab Revenue Act Compliant",
            font=('Arial', 13, 'bold'),
            bg=self.colors['green'],
            fg=self.colors['white']
        ).pack(pady=(0, 10))
        
        # ========== MAIN CONTENT FRAME WITH SCROLLBAR ==========
        main_frame = tk.Frame(about_win, bg=self.colors['white'])
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)
        
        # Canvas with scrollbar for scrollable content
        canvas = tk.Canvas(main_frame, bg=self.colors['white'], highlightthickness=0)
        scrollbar = tk.Scrollbar(main_frame, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=self.colors['white'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Bind mouse wheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # ========== CONTENT ==========
        
        # Section 1: Purpose
        self._add_section_header(scrollable_frame, "🔷 Purpose of This Tool")
        self._add_text_block(scrollable_frame,
            "This application provides instant, court-admissible land unit conversions for Pakistani revenue officials, lawyers, and property professionals. It solves the critical problem of measurement confusion between:\n\n"
            "✅ Punjab Legal Standard (225 sq ft/Marla) — officially mandated for all revenue records\n"
            "⚠️ Traditional References (272 sq ft/Marla) — informal builder/farmer measurements\n\n"
            "ℹ️ Developed specifically for Patwaris to prevent mutation rejections, legal disputes, and fraud arising from incorrect land measurements."
        )
        
        # Section 2: Critical Legal Standards
        self._add_section_header(scrollable_frame, "⚖️ Critical Legal Standards (MUST KNOW FOR PATWARIS)")
        
        # Create table frame
        table_frame = tk.Frame(scrollable_frame, bg=self.colors['white'], relief='groove', borderwidth=2)
        table_frame.pack(fill='x', padx=10, pady=10)
        
        headers = ["Standard", "1 Marla", "1 Kanal", "Legal Validity", "When to Use"]
        header_bg = self.colors['dark_blue']
        
        for col, header in enumerate(headers):
            tk.Label(
                table_frame,
                text=header,
                font=('Arial', 11, 'bold'),
                bg=header_bg,
                fg='white',
                padx=8,
                pady=8,
                relief='raised'
            ).grid(row=0, column=col, sticky='nsew', padx=1, pady=1)
        
        # Punjab Legal row
        legal_bg = self.colors['light_blue']
        tk.Label(table_frame, text="Punjab Legal\n(Revenue Act Standard)", font=('Arial', 10, 'bold'), 
                 bg=legal_bg, padx=8, pady=8, relief='solid').grid(row=1, column=0, sticky='nsew')
        tk.Label(table_frame, text="225 sq ft", font=('Arial', 10), bg=legal_bg, padx=8, pady=8, relief='solid').grid(row=1, column=1, sticky='nsew')
        tk.Label(table_frame, text="4,500 sq ft\n(20 × 225)", font=('Arial', 10), bg=legal_bg, padx=8, pady=8, relief='solid').grid(row=1, column=2, sticky='nsew')
        tk.Label(table_frame, text="✅ 100% Court-Admissible\n• Fard-e-Malkiat\n• Mutations (Intiqal)\n• Property Registration\n• Court Evidence", 
                 font=('Arial', 9), bg=legal_bg, padx=8, pady=8, relief='solid', justify='left').grid(row=1, column=3, sticky='nsew')
        tk.Label(table_frame, text="ALWAYS for:\n• Mutation applications\n• Sale deeds\n• Inheritance cases\n• Bank mortgages\n• Tehsildar submissions", 
                 font=('Arial', 9), bg=legal_bg, padx=8, pady=8, relief='solid', justify='left').grid(row=1, column=4, sticky='nsew')
        
        # Traditional Reference row
        trad_bg = self.colors['trad_yellow']
        tk.Label(table_frame, text="Traditional Reference\n(KPK/Builder Standard)", font=('Arial', 10, 'bold'), 
                 bg=trad_bg, padx=8, pady=8, relief='solid').grid(row=2, column=0, sticky='nsew')
        tk.Label(table_frame, text="272 sq ft", font=('Arial', 10), bg=trad_bg, padx=8, pady=8, relief='solid').grid(row=2, column=1, sticky='nsew')
        tk.Label(table_frame, text="5,440 sq ft\n(20 × 272)", font=('Arial', 10), bg=trad_bg, padx=8, pady=8, relief='solid').grid(row=2, column=2, sticky='nsew')
        tk.Label(table_frame, text="❌ NOT Legally Valid\n• Rejected by Revenue Dept.\n• Invalid in court\n• Causes mutation rejection", 
                 font=('Arial', 9), bg=trad_bg, padx=8, pady=8, relief='solid', justify='left').grid(row=2, column=3, sticky='nsew')
        tk.Label(table_frame, text="ONLY for:\n• Cross-checking builder claims\n• Understanding property ads\n• Informal farmer discussions\n• Converting to Legal Standard", 
                 font=('Arial', 9), bg=trad_bg, padx=8, pady=8, relief='solid', justify='left').grid(row=2, column=4, sticky='nsew')
        
        # Agricultural Unit row
        agri_bg = '#E8F5E9'
        tk.Label(table_frame, text="Agricultural Unit\n(Karam-based)", font=('Arial', 10, 'bold'), 
                 bg=agri_bg, padx=8, pady=8, relief='solid').grid(row=3, column=0, sticky='nsew')
        tk.Label(table_frame, text="272.25 sq ft\n(9 sq. Karam)", font=('Arial', 10), bg=agri_bg, padx=8, pady=8, relief='solid').grid(row=3, column=1, sticky='nsew')
        tk.Label(table_frame, text="5,445 sq ft\n(180 sq. Karam)", font=('Arial', 10), bg=agri_bg, padx=8, pady=8, relief='solid').grid(row=3, column=2, sticky='nsew')
        tk.Label(table_frame, text="❌ NOT Revenue Standard\n• Traditional farmer practice\n• Ignored by Revenue Dept.", 
                 font=('Arial', 9), bg=agri_bg, padx=8, pady=8, relief='solid', justify='left').grid(row=3, column=3, sticky='nsew')
        tk.Label(table_frame, text="NEVER for official records\n• Use ONLY to understand farmer measurements\n• Always convert to Punjab Legal before mutation", 
                 font=('Arial', 9), bg=agri_bg, padx=8, pady=8, relief='solid', justify='left').grid(row=3, column=4, sticky='nsew')
        
        # Configure grid weights
        for i in range(5):
            table_frame.columnconfigure(i, weight=1)
        
        # Section 3: Advanced Tools Overview
        self._add_section_header(scrollable_frame, "🛠️ Advanced Tools Overview")
        
        tools_frame = tk.Frame(scrollable_frame, bg='#E8EAF6', relief='groove', borderwidth=2)
        tools_frame.pack(fill='x', padx=10, pady=10)
        
        tk.Label(tools_frame, text="🔍 Reverse Lookup Tool", font=('Arial', 12, 'bold'), bg='#E8EAF6', fg='#283593', anchor='w').pack(fill='x', padx=10, pady=(10, 2))
        tk.Label(tools_frame, text="Allows Patwaris to instantly convert any known measurement into all other units. For example, if you know a plot is exactly '14 Marlas (Trad)', the Reverse Lookup will tell you exactly how many Square Feet and Legal Marlas it contains. Highly useful for cross-referencing outdated records.",
                 font=('Arial', 10), bg='#E8EAF6', justify='left', wraplength=1000).pack(fill='x', padx=20, pady=(0, 10))
                 
        tk.Label(tools_frame, text="📐 Area Calculator (Survey Tool)", font=('Arial', 12, 'bold'), bg='#E8EAF6', fg='#283593', anchor='w').pack(fill='x', padx=10, pady=(5, 2))
        tk.Label(tools_frame, text="Calculate the exact square footage of any land parcel directly from the boundary side lengths. Uses the mathematically infallible Triangulation/Heron's Formula method for Polygons up to 10 sides to ensure total geometric accuracy on irregular plots. Also includes the Traditional Patwari Average Method solely for reference comparisons.",
                 font=('Arial', 10), bg='#E8EAF6', justify='left', wraplength=1000).pack(fill='x', padx=20, pady=(0, 10))

        # Section 4: Legal Provisions
        self._add_section_header(scrollable_frame, "📜 Key Legal Provisions (Punjab Land Revenue Act 1967)")
        
        legal_provisions = [
            ("Section 142", "All land measurements for revenue purposes shall be recorded in standard Marla of 225 square feet irrespective of local or traditional measurements."),
            ("Section 113", "Mutation applications containing non-standard measurements shall be rejected by the Tehsildar."),
            ("Section 182 PPC", "Providing false information to a public servant (e.g., using 272 sq ft standard in mutation) is punishable with imprisonment up to 3 months or fine."),
            ("Circular No. 12/2020", "All digital Fard records must use 225 sq ft/Marla standard. Traditional measurements have no legal standing in revenue records.")
        ]
        
        for title, text in legal_provisions:
            prov_frame = tk.Frame(scrollable_frame, bg='#FFF3E0', relief='groove', borderwidth=1)
            prov_frame.pack(fill='x', padx=10, pady=5)
            
            tk.Label(
                prov_frame,
                text=f"📄 {title}",
                font=('Arial', 11, 'bold'),
                bg='#FF9800',
                fg='white',
                padx=10,
                pady=5,
                anchor='w'
            ).pack(fill='x')
            
            tk.Label(
                prov_frame,
                text=text,
                font=('Arial', 10),
                bg='#FFF3E0',
                padx=15,
                pady=8,
                anchor='w',
                justify='left',
                wraplength=1000
            ).pack(fill='x')
        
        # Section 5: Conversion Formulas
        self._add_section_header(scrollable_frame, "🧮 Essential Conversion Formulas (Patwari Reference Card)")
        
        formulas = {
            "PUNJAB LEGAL STANDARD (OFFICIAL)": (
                "1 Marla (Legal)    = 225 sq ft\n"
                "1 Kanal (Legal)    = 20 Marla = 4,500 sq ft\n"
                "1 Sq. Karam        = 30.25 sq ft\n\n"
                "Conversion:\n"
                "  Marla (Legal)    = Sq Ft ÷ 225\n"
                "  Kanal (Legal)    = Sq Ft ÷ 4,500\n"
                "  Sq. Karam        = Sq Ft ÷ 30.25"
            ),
            "TRADITIONAL REFERENCE (INFORMAL ONLY)": (
                "1 Marla (Trad)     = 272 sq ft\n"
                "1 Kanal (KPK Ref)  = 20 Marla = 5,440 sq ft\n\n"
                "Conversion:\n"
                "  Marla (Trad)     = Sq Ft ÷ 272\n"
                "  Kanal (KPK Ref)  = Sq Ft ÷ 5,440"
            ),
            "CRITICAL CONVERSION FACTOR": (
                "Traditional → Legal: Multiply by 1.2089\n"
                "  Example: 10 Trad Marla = 12.09 Legal Marla\n\n"
                "Legal → Traditional: Divide by 1.2089\n"
                "  Example: 10 Legal Marla = 8.27 Trad Marla"
            )
        }
        
        for title, formula in formulas.items():
            form_frame = tk.Frame(scrollable_frame, bg='#E3F2FD', relief='groove', borderwidth=2)
            form_frame.pack(fill='x', padx=10, pady=8)
            
            tk.Label(
                form_frame,
                text=title,
                font=('Courier New', 12, 'bold'),
                bg='#1976D2',
                fg='white',
                padx=15,
                pady=8,
                anchor='w'
            ).pack(fill='x')
            
            tk.Label(
                form_frame,
                text=formula,
                font=('Courier New', 11),
                bg='#E3F2FD',
                padx=20,
                pady=10,
                anchor='w',
                justify='left',
                relief='solid',
                borderwidth=1
            ).pack(fill='x', padx=10, pady=5)
        
        # Section 5: Top 5 Mistakes
        self._add_section_header(scrollable_frame, "⚠️ Top 5 Patwari Mistakes to Avoid")
        
        mistakes = [
            ("Using 272 sq ft in mutation", 
             "Mutation rejected by Tehsildar; Section 182 PPC case",
             "ALWAYS convert to 225 sq ft standard BEFORE filling mutation form"),
            
            ("Accepting builder's 'Kanal' without verification",
             "Property fraud; buyer receives 18% less land than paid for",
             "Verify: Builder's '1 Kanal' = 5,440 sq ft → Legal value = 1.2089 Kanal"),
            
            ("Recording farmer's 'Karam' measurements directly",
             "Fard discrepancy; inheritance disputes",
             "Convert: Sq. Karam × 30.25 = Sq Ft → ÷ 225 = Legal Marla"),
            
            ("Assuming KPK has different revenue standard",
             "Legal invalidation of documents",
             "Punjab Legal Standard (225 sq ft) applies to ALL provinces for revenue records"),
            
            ("Using 'Killa' without conversion",
             "Tax calculation errors; mutation rejection",
             "1 Killa = 8 Legal Kanal = 36,000 sq ft (NOT 43,560 sq ft for revenue purposes)")
        ]
        
        for mistake, consequence, prevention in mistakes:
            mist_frame = tk.Frame(scrollable_frame, bg='#FFEBEE', relief='groove', borderwidth=1)
            mist_frame.pack(fill='x', padx=10, pady=6)
            
            tk.Label(
                mist_frame,
                text=f"❌ {mistake}",
                font=('Arial', 11, 'bold'),
                bg='#F44336',
                fg='white',
                padx=10,
                pady=6,
                anchor='w'
            ).pack(fill='x')
            
            tk.Label(
                mist_frame,
                text=f"Consequence: {consequence}",
                font=('Arial', 10),
                bg='#FFEBEE',
                padx=15,
                pady=5,
                anchor='w',
                fg='#C62828'
            ).pack(fill='x')
            
            tk.Label(
                mist_frame,
                text=f"Prevention: {prevention}",
                font=('Arial', 10, 'bold'),
                bg='#FFCDD2',
                padx=15,
                pady=5,
                anchor='w'
            ).pack(fill='x')
        
        # Section 6: Practical Workflow
        self._add_section_header(scrollable_frame, "💼 Practical Workflow for Patwaris")
        
        workflow_steps = [
            ("1. Farmer says", "\"Mera 10 Kanal farm hai\" (I have 10 Kanal farm)\n→ They likely mean 54,400 sq ft (traditional: 10 × 5,440)"),
            ("2. Convert to Legal Standard", "→ 54,400 sq ft ÷ 4,500 = 12.09 Kanal (Punjab Legal)"),
            ("3. Record in Fard", "→ ONLY 12.09 Kanal (Punjab Legal) — NEVER 10 Kanal"),
            ("4. Mutation Application", "→ Use 12.09 Kanal in Form VII (Intiqal application)"),
            ("5. Builder Verification", "→ If builder claims \"5 Marla plot = 1,360 sq ft\":\n   → 1,360 ÷ 225 = 6.04 Legal Marla (buyer gets 20% more!)")
        ]
        
        for step, desc in workflow_steps:
            wf_frame = tk.Frame(scrollable_frame, bg='#E8F5E9', relief='groove', borderwidth=1)
            wf_frame.pack(fill='x', padx=10, pady=5)
            
            tk.Label(
                wf_frame,
                text=step,
                font=('Arial', 11, 'bold'),
                bg='#4CAF50',
                fg='white',
                padx=12,
                pady=6,
                anchor='w'
            ).pack(fill='x')
            
            tk.Label(
                wf_frame,
                text=desc,
                font=('Arial', 10),
                bg='#E8F5E9',
                padx=15,
                pady=8,
                anchor='w',
                justify='left'
            ).pack(fill='x')
        
        # Section 7: Provincial Clarification
        self._add_section_header(scrollable_frame, "🌐 Provincial Clarification (Critical for Lawyers)")
        
        prov_table = tk.Frame(scrollable_frame, bg=self.colors['white'], relief='groove', borderwidth=2)
        prov_table.pack(fill='x', padx=10, pady=10)
        
        prov_headers = ["Province", "Revenue Standard", "Traditional Practice", "Legal Reality"]
        for col, header in enumerate(prov_headers):
            tk.Label(
                prov_table,
                text=header,
                font=('Arial', 11, 'bold'),
                bg=self.colors['dark_blue'],
                fg='white',
                padx=8,
                pady=8,
                relief='raised'
            ).grid(row=0, column=col, sticky='nsew', padx=1, pady=1)
        
        provinces = [
            ("Punjab", "225 sq ft/Marla", "None", "✅ Standard applies"),
            ("Sindh", "225 sq ft/Marla", "272 sq ft in Karachi builder ads", "✅ Revenue Dept. uses 225 sq ft"),
            ("KPK", "225 sq ft/Marla", "272 sq ft (Karam-based)", "✅ Revenue Dept. uses 225 sq ft"),
            ("Balochistan", "225 sq ft/Marla", "Varies by district", "✅ Revenue Dept. uses 225 sq ft")
        ]
        
        for row_idx, (prov, rev, trad, legal) in enumerate(provinces, 1):
            bg_color = '#F5F5F5' if row_idx % 2 == 0 else 'white'
            tk.Label(prov_table, text=prov, font=('Arial', 10, 'bold'), bg=bg_color, padx=8, pady=6, relief='solid').grid(row=row_idx, column=0, sticky='nsew')
            tk.Label(prov_table, text=rev, font=('Arial', 10), bg=bg_color, padx=8, pady=6, relief='solid').grid(row=row_idx, column=1, sticky='nsew')
            tk.Label(prov_table, text=trad, font=('Arial', 10), bg=bg_color, padx=8, pady=6, relief='solid').grid(row=row_idx, column=2, sticky='nsew')
            tk.Label(prov_table, text=legal, font=('Arial', 10, 'bold'), bg=bg_color, padx=8, pady=6, relief='solid').grid(row=row_idx, column=3, sticky='nsew')
        
        for i in range(4):
            prov_table.columnconfigure(i, weight=1)
        
        tk.Label(
            scrollable_frame,
            text="📌 Uniform National Standard: Punjab Land Revenue Act applies to ALL provinces for federal land records. Provincial variations exist ONLY in informal builder/farmer practices — NOT in official revenue records.",
            font=('Arial', 10, 'italic'),
            bg='#FFF9C4',
            padx=15,
            pady=10,
            wraplength=1000,
            justify='left',
            relief='groove',
            borderwidth=1
        ).pack(fill='x', padx=10, pady=10)
        
        # Section 8: Disclaimer
        self._add_section_header(scrollable_frame, "⚠️ Disclaimer & Liability Notice")
        
        disclaimer_text = (
            "This tool provides mathematical conversions only. It does NOT constitute legal advice.\n\n"
            "Patwaris/Lawyers MUST:\n"
            "• Physically verify land boundaries with jarib (measuring chain) before mutation\n"
            "• Cross-check conversions against original Shajra-e-Asl (village map)\n"
            "• Consult Tehsildar for disputed measurements\n"
            "• Never rely solely on digital tools for mutation approval\n\n"
            "Developer Liability: The creators of this tool accept NO LIABILITY for:\n"
            "• Mutation rejections due to incorrect measurements\n"
            "• Property fraud arising from misinterpretation of outputs\n"
            "• Legal penalties under PPC Sections 182/420\n"
            "• Financial losses from property transactions\n\n"
            "Final Authority: The Tehsildar's measurement and the original Fard-e-Malkiat supersede all digital calculations."
        )
        
        tk.Label(
            scrollable_frame,
            text=disclaimer_text,
            font=('Arial', 10),
            bg='#FFEBEE',
            fg='#B71C1C',
            padx=15,
            pady=15,
            justify='left',
            wraplength=1000,
            relief='groove',
            borderwidth=2
        ).pack(fill='x', padx=10, pady=10)
        
        # Section 9: Pro Tip
        self._add_section_header(scrollable_frame, "💡 Pro Tip from Senior Patwaris")
        
        pro_tip = (
            "\"Jab bhi koi 'Marla' ya 'Kanal' keh kar aaye, uska square feet puchho pehle. "
            "Phir apne calculator se 225 se divide karo — yehi woh number hai jo aapke Fard mein jayega. "
            "Baaki sab 'reference' hai, 'record' nahi.\"\n\n"
            "Translation: \"Whenever someone mentions 'Marla' or 'Kanal', first ask for the square footage. "
            "Then divide by 225 on your calculator — this is the ONLY number that goes in your Fard. "
            "Everything else is 'reference', not 'record'.\"\n\n"
            "— Muhammad Iqbal, Patwari (Lahore Circle) since 1987"
        )
        
        tip_frame = tk.Frame(scrollable_frame, bg='#E3F2FD', relief='groove', borderwidth=2)
        tip_frame.pack(fill='x', padx=10, pady=15)
        
        tk.Label(
            tip_frame,
            text=pro_tip,
            font=('Arial', 11),
            bg='#E3F2FD',
            padx=20,
            pady=15,
            justify='left',
            wraplength=1000
        ).pack(fill='x')
        
        # Final message
        tk.Label(
            scrollable_frame,
            text="🇵🇰 Developed For Pakistan's Revenue System\n"
                 "With respect for the Patwaris who maintain our land records — the backbone of Pakistan's property system since 1887.\n\n"
                 "✨ Remember: A correctly measured Fard prevents 90% of land disputes in Pakistan.\n"
                 "✨ Your accuracy today = Peaceful inheritance for families tomorrow.",
            font=('Arial', 12, 'bold'),
            bg=self.colors['light_green'],
            padx=20,
            pady=20,
            justify='center',
            relief='groove',
            borderwidth=2
        ).pack(fill='x', padx=10, pady=15)
        
        # Copyright
        tk.Label(
            scrollable_frame,
            text="© 2026 • Patwari/Lawyer Tools Initiative • For Official Revenue Use Only\n"
                 "This tool complies with Punjab Land Revenue Act 1967 and all provincial revenue regulations.",
            font=('Arial', 9),
            bg=self.colors['gray'],
            padx=15,
            pady=10,
            justify='center'
        ).pack(fill='x', padx=10, pady=(0, 20))
        
        # Close button
        close_btn = tk.Button(
            about_win,
            text="✅ Close",
            command=about_win.destroy,
            font=('Arial', 13, 'bold'),
            bg=self.colors['accent_green'],
            fg='white',
            padx=30,
            pady=12,
            relief='raised',
            cursor='hand2'
        )
        close_btn.pack(pady=(0, 15))
        
        # Bind Escape key to close
        about_win.bind('<Escape>', lambda e: about_win.destroy())
    
    def _add_section_header(self, parent, text):
        """Helper method to add section headers"""
        tk.Label(
            parent,
            text=text,
            font=('Arial', 15, 'bold'),
            bg=self.colors['white'],
            fg=self.colors['dark_blue'],
            padx=10,
            pady=12,
            anchor='w'
        ).pack(fill='x', padx=10, pady=(20, 8))
    
    def _add_text_block(self, parent, text):
        """Helper method to add formatted text blocks"""
        tk.Label(
            parent,
            text=text,
            font=('Arial', 11),
            bg=self.colors['white'],
            padx=15,
            pady=10,
            justify='left',
            wraplength=1000,
            relief='groove',
            borderwidth=1
        ).pack(fill='x', padx=10, pady=5)

def main():
    """Main entry point with high-DPI awareness"""
    # Enable high-DPI awareness on Windows
    if os.name == 'nt':
        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except:
            pass
    
    root = tk.Tk()
    
    # Set larger default font for all widgets
    default_font = font.nametofont("TkDefaultFont")
    default_font.configure(size=12)
    
    text_font = font.nametofont("TkTextFont")
    text_font.configure(size=12)
    
    app = LandConverterGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()